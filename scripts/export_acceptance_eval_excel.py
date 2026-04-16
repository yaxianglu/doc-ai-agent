#!/usr/bin/env python3
from __future__ import annotations

import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def _json_text(value) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2)


def _safe_sheet_title(title: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in title)
    return cleaned[:31] or "Sheet"


def _set_widths(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if "\n" in value:
                value = max(value.splitlines(), key=len, default="")
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 60)


def _write_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _set_widths(ws)


def _build_item_rows(raw_items: list[dict], scored_items: list[dict], baseline_by_index: dict[int, dict]) -> list[list[object]]:
    scored_by_index = {int(item["index"]): item for item in scored_items}
    rows: list[list[object]] = []
    for raw in raw_items:
        index = int(raw["index"])
        scored = scored_by_index.get(index, {})
        evidence = dict(raw.get("evidence") or {})
        response_meta = dict(evidence.get("response_meta") or {})
        request_understanding = dict(evidence.get("request_understanding") or {})
        processing = dict(raw.get("processing") or {})
        baseline_item = baseline_by_index.get(index, {})
        rows.append(
            [
                index,
                str(raw.get("suite") or scored.get("suite") or ""),
                str(raw.get("category") or scored.get("category") or ""),
                str(raw.get("question") or ""),
                str(raw.get("mode") or ""),
                float(raw.get("seconds") or 0),
                float(scored.get("score") or 0),
                float(baseline_item.get("score") or 0) if baseline_item else "",
                (
                    round(float(scored.get("score") or 0) - float(baseline_item.get("score") or 0), 1)
                    if baseline_item else ""
                ),
                ", ".join(list(scored.get("checks_failed") or [])),
                str(evidence.get("generation_mode") or ""),
                str(response_meta.get("fallback_reason") or ""),
                str(request_understanding.get("fallback_reason") or ""),
                float(response_meta.get("confidence") or 0) if response_meta.get("confidence") not in (None, "") else "",
                ", ".join(list(response_meta.get("source_types") or [])),
                str(processing.get("intent_recognition") or ""),
                str(processing.get("data_query") or ""),
                str(processing.get("retrieval") or ""),
                str(processing.get("answer_generation") or ""),
                str(processing.get("ai_involvement") or ""),
                str(raw.get("answer") or ""),
                _json_text(raw.get("turns")),
                _json_text(raw.get("turn_results")),
                _json_text(evidence),
                _json_text(processing),
            ]
        )
    return rows


def _build_turn_rows(raw_items: list[dict], scored_items: list[dict]) -> list[list[object]]:
    scored_by_index = {int(item["index"]): item for item in scored_items}
    rows: list[list[object]] = []
    for raw in raw_items:
        turn_results = list(raw.get("turn_results") or [])
        if not turn_results:
            continue
        turn_scores = list((scored_by_index.get(int(raw["index"]), {}) or {}).get("turn_scores") or [])
        for turn_no, turn in enumerate(turn_results, start=1):
            scored_turn = turn_scores[turn_no - 1] if turn_no - 1 < len(turn_scores) else {}
            rows.append(
                [
                    int(raw["index"]),
                    turn_no,
                    str(raw.get("category") or ""),
                    str(raw.get("suite") or ""),
                    str(turn.get("question") or ""),
                    str(turn.get("mode") or ""),
                    float(turn.get("seconds") or 0),
                    float(scored_turn.get("score") or 0),
                    ", ".join(list(scored_turn.get("checks_failed") or [])),
                    str(turn.get("answer") or ""),
                ]
            )
    return rows


def _build_summary_rows(scored: dict, raw_items: list[dict]) -> list[list[object]]:
    summary = dict(scored.get("summary") or {})
    suite_scores = dict(summary.get("suite_scores") or {})
    category_scores = dict(summary.get("category_scores") or {})
    rows: list[list[object]] = [
        ["generated_at", str(scored.get("generated_at") or "")],
        ["count", int(summary.get("count") or 0)],
        ["average_score", float(summary.get("average_score") or 0)],
        ["low_score_count", int(summary.get("low_score_count") or 0)],
        ["raw_item_count", len(raw_items)],
    ]
    for name, score in category_scores.items():
        rows.append([f"category::{name}", score])
    for name, score in suite_scores.items():
        rows.append([f"suite::{name}", score])
    return rows


def _build_check_rows(scored_items: list[dict]) -> list[list[object]]:
    counter: Counter[str] = Counter()
    for item in scored_items:
        for check in list(item.get("checks_failed") or []):
            counter[str(check)] += 1
    return [[name, count] for name, count in counter.most_common()]


def main() -> int:
    if len(sys.argv) not in {4, 5}:
        raise SystemExit("usage: export_acceptance_eval_excel.py RAW_JSON SCORED_JSON OUTPUT_XLSX [BASELINE_SCORED_JSON]")

    raw_path = Path(sys.argv[1])
    scored_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])
    baseline_path = Path(sys.argv[4]) if len(sys.argv) == 5 else None

    raw_items = json.loads(raw_path.read_text(encoding="utf-8"))
    scored = json.loads(scored_path.read_text(encoding="utf-8"))
    scored_items = list(scored.get("items") or [])
    baseline_by_index = {}
    if baseline_path and baseline_path.exists():
        baseline = json.loads(baseline_path.read_text(encoding="utf-8"))
        baseline_by_index = {int(item["index"]): item for item in list(baseline.get("items") or [])}

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    _write_sheet(summary_ws, ["metric", "value"], _build_summary_rows(scored, raw_items))

    items_ws = workbook.create_sheet("Items")
    _write_sheet(
        items_ws,
        [
            "index",
            "suite",
            "category",
            "question",
            "mode",
            "seconds",
            "score",
            "baseline_score",
            "delta_vs_baseline",
            "checks_failed",
            "generation_mode",
            "response_fallback_reason",
            "request_fallback_reason",
            "response_confidence",
            "source_types",
            "processing_intent_recognition",
            "processing_data_query",
            "processing_retrieval",
            "processing_answer_generation",
            "processing_ai_involvement",
            "answer",
            "turns_json",
            "turn_results_json",
            "evidence_json",
            "processing_json",
        ],
        _build_item_rows(raw_items, scored_items, baseline_by_index),
    )

    turns_ws = workbook.create_sheet("Turns")
    _write_sheet(
        turns_ws,
        ["index", "turn_no", "category", "suite", "question", "mode", "seconds", "score", "checks_failed", "answer"],
        _build_turn_rows(raw_items, scored_items),
    )

    low_ws = workbook.create_sheet("LowScore")
    low_rows = [
        [
            int(item["index"]),
            str(item.get("suite") or ""),
            str(item.get("category") or ""),
            float(item.get("score") or 0),
            ", ".join(list(item.get("checks_failed") or [])),
            str(item.get("question") or ""),
            str(item.get("answer") or ""),
        ]
        for item in scored_items
        if float(item.get("score") or 0) < 7.0
    ]
    _write_sheet(low_ws, ["index", "suite", "category", "score", "checks_failed", "question", "answer"], low_rows)

    checks_ws = workbook.create_sheet("Checks")
    _write_sheet(checks_ws, ["check", "count"], _build_check_rows(scored_items))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(json.dumps({"output_xlsx": str(output_path)}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
