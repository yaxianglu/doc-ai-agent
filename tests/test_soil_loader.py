import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from doc_ai_agent.soil_loader import iter_rows


class SoilLoaderTests(unittest.TestCase):
    def test_iter_rows_reads_direct_region_and_latlon_fields(self):
        headers = [
            "id",
            "sn",
            "gatewayid",
            "sensorid",
            "unitid",
            "time",
            "water20cm",
            "water40cm",
            "water60cm",
            "water80cm",
            "t20cm",
            "t40cm",
            "t60cm",
            "t80cm",
            "water20cmfieldstate",
            "water40cmfieldstate",
            "water60cmfieldstate",
            "water80cmfieldstate",
            "t20cmfieldstate",
            "t40cmfieldstate",
            "t60cmfieldstate",
            "t80cmfieldstate",
            "create_time",
            "lat",
            "lon",
            "city",
            "county",
        ]
        row = [
            "record-1",
            "SNS001",
            "gw-1",
            "sensor-1",
            "unit-1",
            "45410.0",
            "48.5",
            "50.1",
            "51.2",
            "52.3",
            "19.8",
            "19.1",
            "18.7",
            "18.1",
            "ok",
            "ok",
            "ok",
            "ok",
            "ok",
            "ok",
            "ok",
            "ok",
            "45410.0",
            "32.328056",
            "120.974167",
            "南通市",
            "如东县",
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "soil.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet.append(headers)
            sheet.append(row)
            workbook.save(path)

            rows = list(iter_rows(str(path), "batch-1"))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["city_name"], "南通市")
        self.assertEqual(rows[0]["county_name"], "如东县")
        self.assertEqual(rows[0]["latitude"], 32.328056)
        self.assertEqual(rows[0]["longitude"], 120.974167)


if __name__ == "__main__":
    unittest.main()
